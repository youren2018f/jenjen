#清洗資料
import re
import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl.writer.excel import save_virtual_workbook
import jieba
import openpyxl
from datetime import date

def generate_vectors(vec_a, vec_b):
    check_dict = dict.fromkeys(set(vec_a) | set(vec_b), 0) #產生2個向量的交集的字典，且設定value為0
    #比對check_dict，如果有存在則+1
    import copy
    c1_copy = copy.deepcopy(check_dict)
    for a in vec_a:
        for key in c1_copy.keys():
            if a == key:
                c1_copy[key] += 1
    consine_list_1 = [value for value in c1_copy.values()] #轉換為cosine_list
    c2_copy = copy.deepcopy(check_dict)
    for b in vec_b:
        for key in c2_copy.keys():
            if b == key:
                c2_copy[key] += 1
    consine_list_2 = [value for value in c2_copy.values()] #轉換為cosine_list
    return (consine_list_1, consine_list_2)

#計算Cosine Similarity
def cosine_similarity(vec_ab):
    vec_a = vec_ab[0]
    vec_b = vec_ab[-1]
    # Dot and norm
    dot = sum(a*b for a, b in zip(vec_a, vec_b))
    norm_a = sum(a*a for a in vec_a) ** 0.5
    norm_b = sum(b*b for b in vec_b) ** 0.5

    # Cosine similarity
    cos_sim = dot / (norm_a*norm_b)
    return cos_sim


def rename(inds):
    shortest_word = min(inds, key = len)
    re_name_list = []
    for word in shortest_word:
        matrix = []
        for ind in inds:
            if word in ind:
                matrix.append(1)
            else:
                matrix.append(0)
        if all(matrix):
            re_name_list.append(word)
    name = "".join(re_name_list)
    my_str = f"改為名稱:{name}"
    st.write(my_str)
    name_dict = {}
    name_dict[name] = inds
    for key,values in name_dict.items():
        for value in values:
            try:
                df["插播名稱"] = df["插播名稱"].replace(value, key)
            except:
                pass

def compute(df):
    for i in range(len(df)):       
        str = df["插播名稱"][i]
        df["插播名稱"][i] = re.sub('\d{7}', '', str)

    duplicate = [] #為了移除（1,3,4)之後的(3,4)
    for index_x in range(len(df)):
        same=[]
        if index_x not in set(duplicate):
            for index_y in range(index_x + 1, len(df)):  #避免出現(3,2)(2,3)及(3,3)的狀況
                #分詞並去除符號
                filter_elems = [" ", "(", ")", "_", "-", "、", "，", "《", "》"]
                vec_a = jieba.lcut(df['插播名稱'][index_x])
                for filter_elem in filter_elems:                
                    vec_a = [x for x in vec_a if x!= filter_elem]
                vec_b = jieba.lcut(df['插播名稱'][index_y])
                for filter_elem in filter_elems:                
                    vec_b = [x for x in vec_b if x!= filter_elem]

                #計算餘弦相似度
                vec_ab = generate_vectors(vec_a, vec_b)

                similarity = cosine_similarity(vec_ab)

                if similarity >= 0.67:
                    #print(df['插播名稱'][index_x], df['插播名稱'][index_y])
                    same.append(df['插播名稱'][index_x])
                    same.append(df['插播名稱'][index_y])
                    duplicate.append(index_x)
                    duplicate.append(index_y)

            if same:
                if len(set(same)) > 1:
                    st.write(set(same))
                    rename(set(same))
            #print(same)
                

    # seg_list = jieba.lcut(sentence)


    my_dict = {}
    for index, row in df.iterrows():
        if row['插播名稱'] not in my_dict.keys():
            my_dict[row['插播名稱']] = row['播放次數']
        else:        my_dict[row['插播名稱']] += row['播放次數']

    df2 = pd.DataFrame(list(my_dict.items()),columns = ['插播名稱','播放次數']) 
    return df2



st.title('''
以餘弦相似度(Cosine Similarity)計算插播統計表相似度
''')


input_month = st.text_input('請輸入月份數字，預設值為當下時間的上一個月', str(date.today().month-1))

st.header('上傳區')




df_all = pd.DataFrame()
uploaded_files = st.file_uploader("請依序上傳(A-E)插播統計表xlsx檔, 檔名內要有(A-E)類別才會有對應的值", type = ".xlsx", accept_multiple_files=True)


if st.button('點擊開始運作...'):

    for uploaded_file in uploaded_files:
        
        df = pd.read_excel(uploaded_file, header=0)

        if "A" in uploaded_file.name:   
            st.title("在A類中合併的項目:")       
            df2 = compute(df)
            df2["類別"] = "A"
        elif "B" in uploaded_file.name:
            st.title("在B類中合併的項目:")   
            df2 = compute(df)
            df2["類別"] = "B"
            
        elif "C" in uploaded_file.name:
            st.title("在C類中合併的項目:")   
            df2 = compute(df)
            df2["類別"] = "C"
            
        elif "D" in uploaded_file.name:
            st.title("在D類中合併的項目:")   
            df2 = compute(df)
            df2["類別"] = "D"
       
        elif "E" in uploaded_file.name:
            st.title("在E類中合併的項目:")   
            df2 = compute(df)
            df2["類別"] = "E"
        
        else:
            st.title("在未知類中合併的項目:")   
            df2 = compute(df)
            df2["類別"] = "unknow"
            
     
 
        df_all = pd.concat([df_all, df2], ignore_index=True)



    df_all["單位"] = "臺東分臺"
    df_all["日期"] = str(input_month) + "月"
    df_all = df_all.reindex(columns = ["單位", "日期", "插播名稱", "類別", "播放次數"])
    wb = openpyxl.load_workbook(r"blank.xlsx")
    #指定那一個worksheet
    ws =wb['8.臺東分台']
    ws.cell(row=4, column=2).value = str(input_month) + "月"
    #開始疊代
 

    i = 5 #從第幾列開始疊代
    for ind in df_all.index:
        ws.cell(row=i, column=1).value = df_all.loc[ind, "單位"]
        ws.cell(row=i, column=2).value = df_all.loc[ind, "日期"]
        ws.cell(row=i, column=5).value = df_all.loc[ind, "插播名稱"]
        ws.cell(row=i, column=6).value = df_all.loc[ind, "類別"]
        ws.cell(row=i, column=7).value = df_all.loc[ind, "播放次數"]
        i = i + 1


    data = BytesIO(save_virtual_workbook(wb))
    
    st.header('下載區')
    st.download_button("下載檔案",
        data=data,
        mime='xlsx',
        file_name="再撐一下就下班.xlsx")
    st.write("按下下載檔案後，頁面會重載，這個是模組的問題，尚無法解決。")













# if uploaded_file is not None:

#     df = pd.read_excel(source_file, header=0)
#     df2 = compute(df)
    
    
#     st.write("====================================================")
#     st.write("檔案下載預覽")
#     st.table(df2)
    
#     wb = openpyxl.load_workbook(r"blank.xlsx")
#     #指定那一個worksheet
#     ws =wb['統計結果']
#     #開始疊代
#     i = 2 #從第幾列開始
#     for ind in df2.index:
#         ws.cell(row=i, column=1).value = df2.loc[ind, "插播名稱"]
#         ws.cell(row=i, column=2).value = df2.loc[ind, "播放次數"]
#         i = i + 1
#     data = BytesIO(save_virtual_workbook(wb))
#     st.header('下載區')
#     st.download_button("下載檔案",
#         data=data,
#         mime='xlsx',
#         file_name="插播統計表_修改過.xlsx")

